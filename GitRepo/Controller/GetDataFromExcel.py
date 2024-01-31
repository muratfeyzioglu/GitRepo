import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
class GetDataFromExcel:

    excel_file = "Excel/File/Directory/Filename.xlsx"

    value1 = ""
    value2 = ""

    num_rows = 0
    num_cols = 0



    def getData(self):
        # Read the Excel file into a pandas DataFrame
        df = pd.read_excel(self.excel_file)
        return df

    """
    The fillDatas method assigns the data from the Excel sheet to the relevant variables.
     The row number is obtained from the rowKey field and added to the variable to be used.
    """
    def fillDatas(self, rowKey):
        df = self.getData()

        num_rows, num_cols = df.shape
        print("Number of rows:", num_rows)
        print("Number of columns:", num_cols)

        j = 0
        self.value1 = df.iloc[rowKey, j]
        j += 1
        self.value2 = df.iloc[rowKey, j]
        j += 1


    """changeColor to  selected Cell"""
    def changeColor(self, row, column):
        # Load the Excel workbook
        df = self.getData()

        workbook = openpyxl.load_workbook(self.excel_file)
        sheet = workbook.active
        cell = sheet.cell(row=row+2, column=column+1)

        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                fill_type='solid')  # Change 'FF0000' to the desired color code
        workbook.save(self.excel_file)
